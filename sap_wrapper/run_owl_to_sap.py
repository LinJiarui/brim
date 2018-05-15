import xlrd
import xml.etree.ElementTree as ET
import xml.dom.minidom as minidom
import openpyxl
from dict_sap_to_owl_v2 import dict_sap_to_owl

#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-##-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#
# string constant
#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-##-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#
ns_brim = "http://eil.stanford.edu/ontologies/brim"
ns_base = "http://eil.stanford.edu/ontologies/brim"
ns_rdf = "http://www.w3.org/1999/02/22-rdf-syntax-ns"
ns_owl = "http://www.w3.org/2002/07/owl"
ns_xml = "http://www.w3.org/XML/1998/namespace"
ns_xsd = "http://www.w3.org/2001/XMLSchema"
ns_rdfs = "http://www.w3.org/2000/01/rdf-schema"

ns_base_s = ns_base+"#"
ns_brim_s = ns_brim+"#"
ns_rdf_s = ns_rdf+'#'
ns_owl_s = ns_owl+'#'
ns_xml_s = ns_xml+'#'
ns_xsd_s = ns_xsd+'#'
ns_rdfs_s = ns_rdfs+'#'

ns_base_p = "{"+ns_base_s+"}"
ns_brim_p = "{"+ns_brim_s+"}"
ns_rdf_p = "{"+ns_rdf_s+"}"
ns_owl_p = "{"+ns_owl_s+"}"
ns_xml_p = "{"+ns_xml_s+"}"
ns_xsd_p = "{"+ns_xsd_s+"}"
ns_rdfs_p = "{"+ns_rdfs_s+"}"


templateFileName = 'New_TRB4_cleaned_modal_only_template.xlsx'
inputFileName = 'testoutput.owl'
outputFileName = 'testoutput.xlsx'

#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#
# global variable
#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#

dataEntity = {}
duplicationAllowList = ['hasBridgeLayoutHorizontalSegment', 
                        'hasBridgeLayoutVerticalSegment',
                        'hasFrequencyFunctionValue',
                        'hasPeriodFunctionValue',
                        'hasTimeSeriesFunctionValue',
                        'hasLaneSegment',
                        'hasAreaJointOffset',
                        'hasSubSectionPlate']
varForAllowingDuplication = [0]


#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#
# functions
#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#

def ScanAttributes(dictionary, sheet):
	ncol = sheet.max_column
	columnList = []
	for i in range(0, ncol):
		columnList.append(sheet.cell(row = 2, column = i+1).value)
	columnDict = {}
	for key in dictionary:
		if key in columnList:
			columnDict[columnList.index(key)] = dictionary[key]
	return columnDict

# def GetParameter(params):
# 	parameter = {}
# 	paramAttr = {}
# 	if params :
# 		for param in params:
# 			paramAttr = GetAttribute(param.items())
# 			paramVal = paramAttr['V']
# 			if 'Type' not in paramAttr:
# 				paramVal = float(paramVal)
# 			parameter[paramAttr['N']] = paramVal
# 	return parameter


# def GetAttribute(items):
# 	attribute = {}
# 	for item in items:
# 		attribute[item[0]] = item[1]
# 	return attribute



# 	for childObj in corrObj.findall('O'):
# 		print childObj



# def GetObjectID(obj):
# 	return obj.attrib["{0}about".format(ns_rdf_p)].split('#')[1]

# def GetClassID(obj):
# 	attrib = obj.findall("{0}type".format(ns_rdf_p))[0].attrib
# 	return attrib["{0}resource".format(ns_rdf_p)].split('#')[1]

# def GetDataProperty(obj):
# 	dataProperty = {}
# 	objectProperty = {}
# 	objID = obj.attrib["{0}about".format(ns_rdf_p)].split('#')[1]
# 	for elem in obj.findall("*"):
# 		if elem.tag == "{0}type".format(ns_rdf_p):
# 			classID = elem.attrib["{0}resource".format(ns_rdf_p)].split('#')[1]
# 		elif elem.text != None:
# 			dataProperty[elem.tag.split('#}')[1]] = elem.text
# 		else:
# 			objectProperty[elem.tag.split('#}')[1]] = elem.attrib["{0}resource".format(ns_rdf_p)].split('#')[1]
# 	return objID, classID, dataProperty, objectProperty



def ParseObject(obj):
	dataProperty = {}
	objectProperty = {}
	objID = obj.attrib["{0}about".format(ns_rdf_p)].split('#')[1]
	for elem in obj.findall("*"):
		if elem.tag == "{0}type".format(ns_rdf_p):
			classID = elem.attrib["{0}resource".format(ns_rdf_p)].split('#')[1]
		elif elem.text != None:
			dataProperty[elem.tag.split('#}')[1]] = elem.text
		else:
			# if objectPropertyKey = objectProperty + '____' + str(int(varForAllowingDuplication[0]))
			if elem.tag.split('#}')[1] in duplicationAllowList:
				objectPropertyKey = elem.tag.split('#}')[1] + '____' + str(int(varForAllowingDuplication[0]))
				varForAllowingDuplication[0] += 1
				objectProperty[objectPropertyKey] = elem.attrib["{0}resource".format(ns_rdf_p)].split('#')[1]
			else:
				objectProperty[elem.tag.split('#}')[1]] = elem.attrib["{0}resource".format(ns_rdf_p)].split('#')[1]

	return objID, classID, dataProperty, objectProperty



def sortingFunction(item):
	if '__' in item:
		return int(item.split('__')[1])



def mapToTable_AreaAutoMeshAssignments(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex):
	invForiegnKey = invForiegnKeyIndex[invForiegnKeyIndex.keys()[0]]
	invForiegnKeyClassID = invForiegnKey[1][0]
	invForiegnKeyObjectAttribute = invForiegnKey[0]
	keys = sorted(dataEntity[invForiegnKeyClassID].keys(), key=sortingFunction)
	rowIdx = 4
	for key in keys:
		if invForiegnKeyObjectAttribute in dataEntity[invForiegnKeyClassID][key][1]:
			objectID = dataEntity[invForiegnKeyClassID][key][1][invForiegnKeyObjectAttribute]
			data = dataEntity[classIDCandidate[0]][objectID][0]

			for i in invForiegnKeyIndex:
				sheet.cell(row = rowIdx, column = i+1).value = key.split('__')[1]

			for i in attributeIndex:
				value = data[attributeIndex[i][0]]
				if attributeIndex[i][1] == bool:
					value = "Yes" if value == 'true' else "No"
				sheet.cell(row = rowIdx, column = i+1).value = value
				sheet.cell(row = rowIdx, column = 2).value = "Points On Edges"
				sheet.cell(row = rowIdx, column = 3).value = "ALL"
			rowIdx = rowIdx + 1

				
def mapToTable_AreaOverwritesJointOffsets(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex):
	invForiegnKey = invForiegnKeyIndex[invForiegnKeyIndex.keys()[0]]
	invForiegnKeyClassID = invForiegnKey[1][0]
	invForiegnKeyObjectAttribute = invForiegnKey[0]
	keys = sorted(dataEntity[invForiegnKeyClassID].keys(), key=sortingFunction)
	rowIdx = 4
	
	for key in keys:


		for attr in dataEntity[invForiegnKeyClassID][key][1]:
			if '____' in attr:
				thisAttr = attr.split('____')[0]
				if thisAttr == invForiegnKeyObjectAttribute:

					objectID = dataEntity[invForiegnKeyClassID][key][1][attr]
					data = dataEntity[classIDCandidate[0]][objectID][0]
					objectData = dataEntity[classIDCandidate[0]][objectID][1]

		# if invForiegnKeyObjectAttribute in dataEntity[invForiegnKeyClassID][key][1]:

					sheet.cell(row = rowIdx, column = 2).value = "Object"
					dataIdx = 0
					# objectID = dataEntity[invForiegnKeyClassID][key][1][invForiegnKeyObjectAttribute]
					# data = dataEntity[classIDCandidate[0]][objectID][0]

					while dataIdx < len(data):
						if (dataIdx) % 4 == 0:
							for i in invForiegnKeyIndex:
								sheet.cell(row = rowIdx, column = i+1).value = key.split('__')[1]
						attr = attributeIndex[dataIdx%4+2][0]+ str(int(dataIdx+1))
						value = data[attr]
						sheet.cell(row = rowIdx, column = dataIdx%4+3).value = value
						if dataIdx + 1 < len(data) and (dataIdx + 1) % 4 == 0:
							rowIdx += 1
						dataIdx +=1
					rowIdx = rowIdx + 1

def mapToTable_MatProp01General(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex):
	rowIdx = 4
	for classID in classIDCandidate:
		for key in dataEntity[classID]:
			data = dataEntity[classID][key][0]
			
			for i in attributeIndex:
				if attributeIndex[i][0] == 'ID':
					value = key.split('__')[1]
				elif attributeIndex[i][0] == 'classID':
					value = classID
				else:
					try:
						value = data[attributeIndex[i][0]]
					except:
						continue
				if attributeIndex[i][1] == bool:
					value = "Yes" if value == 'true' else "No"
				sheet.cell(row = rowIdx, column = i+1).value = value
			if classID in ['Concrete', 'Steel']:
				sheet.cell(row = rowIdx, column = 3).value = 'Isotropic'
			elif classID in ['Rebar', 'Tendon']:
				sheet.cell(row = rowIdx, column = 3).value = 'Uniaxial'
			rowIdx += 1



def mapToTable_Type1(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex):
	invForiegnKey = invForiegnKeyIndex[invForiegnKeyIndex.keys()[0]]
	invForiegnKeyClassID = invForiegnKey[1][0]
	invForiegnKeyObjectAttribute = invForiegnKey[0]
	try:
		keys = sorted(dataEntity[invForiegnKeyClassID].keys(), key=sortingFunction)
	except:
		keys = dataEntity[invForiegnKeyClassID].keys()
	rowIdx = 4
	for key in keys:
		if invForiegnKeyObjectAttribute in dataEntity[invForiegnKeyClassID][key][1]:
			objectID = dataEntity[invForiegnKeyClassID][key][1][invForiegnKeyObjectAttribute]
			data = dataEntity[classIDCandidate[0]][objectID][0]
			objectData = dataEntity[classIDCandidate[0]][objectID][1]

			for i in invForiegnKeyIndex:
				sheet.cell(row = rowIdx, column = i+1).value = key.split('__')[1]

			for i in attributeIndex:
				value = data[attributeIndex[i][0]]
				if attributeIndex[i][1] == bool:
					value = "Yes" if value == 'true' else "No"
				sheet.cell(row = rowIdx, column = i+1).value = value

			for i in foriegnKeyIndex:
				value = objectData[foriegnKeyIndex[i][0]]
				sheet.cell(row = rowIdx, column = i+1).value = value.split('__')[1]

			rowIdx = rowIdx + 1


def mapToTable_Type2(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex):
	rowIdx = 4
	for classID in classIDCandidate:
		for key in dataEntity[classID]:
			data = dataEntity[classID][key][0]
			
			for i in attributeIndex:
				if attributeIndex[i][0] == 'ID':
					value = key.split('__')[1]
				elif attributeIndex[i][0] == 'classID':
					value = classID
				else:
					try:
						value = data[attributeIndex[i][0]]
					except:
						continue
				if attributeIndex[i][1] == bool:
					value = "Yes" if value == 'true' else "No"
				sheet.cell(row = rowIdx, column = i+1).value = value
			rowIdx += 1


def mapToTable_AreaSectionProperties(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex):
	rowIdx = 4
	for classID in classIDCandidate:
		for key in dataEntity[classID]:
			dataProperty = dataEntity[classID][key][0]
			objectProperty = dataEntity[classID][key][1]

			for i in attributeIndex:
				if attributeIndex[i][0] == 'ID':
					value = key.split('__')[1]
				elif attributeIndex[i][0] == 'classID':
					value = classID
				else:
					try:
						value = dataProperty[attributeIndex[i][0]]
					except:
						continue
				if attributeIndex[i][1] == bool:
					value = "Yes" if value == 'true' else "No"
				sheet.cell(row = rowIdx, column = i+1).value = value

			for i in foriegnKeyIndex:
				value = objectProperty[foriegnKeyIndex[i][0]].split('__')[1]
				sheet.cell(row = rowIdx, column = i+1).value = value
			sheet.cell(row = rowIdx, column = 5).value = "Shell-Thin"
			rowIdx += 1



def mapToTable_Type3(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex):
	invForiegnKey = invForiegnKeyIndex[invForiegnKeyIndex.keys()[0]]
	invForiegnKeyClassID = invForiegnKey[1][0]
	invForiegnKeyObjectAttribute = invForiegnKey[0]
	keys = sorted(dataEntity[invForiegnKeyClassID].keys(), key=sortingFunction)
	rowIdx = 4
	
	for key in keys:
		if invForiegnKeyObjectAttribute in dataEntity[invForiegnKeyClassID][key][1]:
			objectID = dataEntity[invForiegnKeyClassID][key][1][invForiegnKeyObjectAttribute]
			data = dataEntity[classIDCandidate[0]][objectID][0]
			for i in invForiegnKeyIndex:
				if '__' in key:
					sheet.cell(row = rowIdx, column = i+1).value = key.split('__')[1]
				else: 
					sheet.cell(row = rowIdx, column = i+1).value = key

			for i in attributeIndex:
				if attributeIndex[i][0] == 'ID':
					value = objectID.split('__')[1]
					
			# 	if attributeIndex[i][1] == bool:
			# 		value = "Yes" if value == 'true' else "No"
				sheet.cell(row = rowIdx, column = i+1).value = value
			rowIdx = rowIdx + 1




def mapToTable_CoordinateSystems(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex):
	invForiegnKey = invForiegnKeyIndex[invForiegnKeyIndex.keys()[0]]
	invForiegnKeyClassID = invForiegnKey[1][0]
	invForiegnKeyObjectAttribute = invForiegnKey[0]
	try:
		keys = sorted(dataEntity[invForiegnKeyClassID].keys(), key=sortingFunction)
	except:
		keys = dataEntity[invForiegnKeyClassID].keys()
	rowIdx = 4
	for key in keys:
		if invForiegnKeyObjectAttribute in dataEntity[invForiegnKeyClassID][key][1]:
			objectID = dataEntity[invForiegnKeyClassID][key][1][invForiegnKeyObjectAttribute]
			data = dataEntity[classIDCandidate[0]][objectID][0]

			for i in invForiegnKeyIndex:
				sheet.cell(row = rowIdx, column = i+1).value = key.split('__')[1]

			for i in attributeIndex:
				value = data[attributeIndex[i][0]]
				if attributeIndex[i][1] == bool:
					value = "Yes" if value == 'true' else "No"
				sheet.cell(row = rowIdx, column = i+1).value = value
			sheet.cell(row = rowIdx, column = 2).value = "Cartesian"
			rowIdx = rowIdx + 1



def mapToTable_ConnectivityArea(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex):
	rowIdx = 4
	classID = classIDCandidate[0]
	keys = sorted(dataEntity[classID].keys(), key=sortingFunction)
	for key in keys:
		data = dataEntity[classID][key][1]
		dataIdx = 0
		dataLength = 0
		while True:
			if 'hasAreaNode'+str(int(dataLength+1)) in data:
				dataLength += 1
			else:
				break
		sheet.cell(row = rowIdx, column = 2).value = dataLength

		while dataIdx < dataLength:
			if (dataIdx) % 4 == 0:
				for i in attributeIndex:
					sheet.cell(row = rowIdx, column = i+1).value = key.split('__')[1]
			attr = foriegnKeyIndex[dataIdx%4+2][0]+ str(int(dataIdx+1))
			value = data[attr]
			sheet.cell(row = rowIdx, column = dataIdx%4+3).value = value.split('__')[1]
			if dataIdx + 1 < dataLength and (dataIdx + 1) % 4 == 0:
				rowIdx += 1
			dataIdx +=1
		rowIdx += 1



def mapToTable_ConnectivityFrame(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex):
	rowIdx = 4
	classID = classIDCandidate[0]
	keys = sorted(dataEntity[classID].keys(), key=sortingFunction)
	for key in keys:
		# data = dataEntity[classID][key][1]
		objectProperty = dataEntity[classID][key][1]

		sheet.cell(row = rowIdx, column = 1).value = key.split('__')[1]
		
		for i in foriegnKeyIndex:
			value = objectProperty[foriegnKeyIndex[i][0]].split('__')[1]
			sheet.cell(row = rowIdx, column = i+1).value = value
		sheet.cell(row = rowIdx, column = 4).value = "No"
		rowIdx += 1


def mapToTable_JointCoordinates(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex):
	invForiegnKey = invForiegnKeyIndex[invForiegnKeyIndex.keys()[0]]
	invForiegnKeyClassID = invForiegnKey[1][0]
	invForiegnKeyObjectAttribute = invForiegnKey[0]
	try:
		keys = sorted(dataEntity[invForiegnKeyClassID].keys(), key=sortingFunction)
	except:
		keys = dataEntity[invForiegnKeyClassID].keys()
	rowIdx = 4
	for key in keys:
		if invForiegnKeyObjectAttribute in dataEntity[invForiegnKeyClassID][key][1]:
			objectID = dataEntity[invForiegnKeyClassID][key][1][invForiegnKeyObjectAttribute]
			data = dataEntity[classIDCandidate[0]][objectID][0]
			objectData = dataEntity[classIDCandidate[0]][objectID][1]

			for i in invForiegnKeyIndex:
				sheet.cell(row = rowIdx, column = i+1).value = key.split('__')[1]

			for i in attributeIndex:
				value = data[attributeIndex[i][0]]
				if attributeIndex[i][1] == bool:
					value = "Yes" if value == 'true' else "No"
				sheet.cell(row = rowIdx, column = i+1).value = value

			for i in foriegnKeyIndex:
				value = objectData[foriegnKeyIndex[i][0]]
				sheet.cell(row = rowIdx, column = i+1).value = value.split('__')[1]
			sheet.cell(row = rowIdx, column = 3).value = 'Cartesian'
			sheet.cell(row = rowIdx, column = 7).value = 'No'
			rowIdx = rowIdx + 1



def mapToTable_Type4(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex):
	invForiegnKey = invForiegnKeyIndex[invForiegnKeyIndex.keys()[0]]
	invForiegnKeyClassID = invForiegnKey[1][0]
	invForiegnKeyObjectAttribute = invForiegnKey[0]
	keys = sorted(dataEntity[invForiegnKeyClassID].keys(), key=sortingFunction)
	rowIdx = 4

	for key in keys:
		if invForiegnKeyObjectAttribute in dataEntity[invForiegnKeyClassID][key][1]:
			objectID = dataEntity[invForiegnKeyClassID][key][1][invForiegnKeyObjectAttribute]
			data = dataEntity[classIDCandidate[0]][objectID][0]

			for i in invForiegnKeyIndex:
				sheet.cell(row = rowIdx, column = i+1).value = key.split('__')[1]

			for i in attributeIndex:
				value = data[attributeIndex[i][0]]
				if attributeIndex[i][1] == bool:
					value = "Yes" if value == 'true' else "No"
				sheet.cell(row = rowIdx, column = i+1).value = value
			rowIdx = rowIdx + 1



def mapToTable_FramePropsGeneral(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex):
	rowIdx = 4
	sapClassIDs = {'LineSectionAngle':'Angle',
	               'LineSectionTShape':'Tee',
	               'LineSectionCShape':'Channel',
	               'LineUserDefinedSection':'SD Section'}
	for classID in classIDCandidate:
		for key in dataEntity[classID]:
			dataProperty = dataEntity[classID][key][0]
			objectProperty = dataEntity[classID][key][1]

			for i in attributeIndex:
				if attributeIndex[i][0] == 'ID':
					value = key.split('__')[1]
				elif attributeIndex[i][0] == 'classID':
					value = sapClassIDs[classID]
				else:
					try:
						value = dataProperty[attributeIndex[i][0]]
					except:
						continue
				if attributeIndex[i][1] == bool:
					value = "Yes" if value == 'true' else "No"
				sheet.cell(row = rowIdx, column = i+1).value = value

			for i in foriegnKeyIndex:
				value = objectProperty[foriegnKeyIndex[i][0]].split('__')[1]
				sheet.cell(row = rowIdx, column = i+1).value = value
			sheet.cell(row = rowIdx, column = 5).value = "Shell-Thin"
			rowIdx += 1


def mapToTable_FrameSectionAssignments(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex):
	rowIdx = 4
	for classID in classIDCandidate:
		keys = sorted(dataEntity[classID].keys(), key=sortingFunction)
		for key in keys:
			dataProperty = dataEntity[classID][key][0]
			objectProperty = dataEntity[classID][key][1]

			for i in attributeIndex:
				if attributeIndex[i][0] == 'ID':
					value = key.split('__')[1]
				elif attributeIndex[i][0] == 'classID':
					value = classID
				else:
					try:
						value = dataProperty[attributeIndex[i][0]]
					except:
						continue
				if attributeIndex[i][1] == bool:
					value = "Yes" if value == 'true' else "No"
				sheet.cell(row = rowIdx, column = i+1).value = value

			for i in foriegnKeyIndex:
				value = objectProperty[foriegnKeyIndex[i][0]].split('__')[1]
				sheet.cell(row = rowIdx, column = i+1).value = value
			# sheet.cell(row = rowIdx, column = 5).value = "Shell-Thin"
			rowIdx += 1


def mapToTable_ShapePlate(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex):
	invForiegnKey = invForiegnKeyIndex[invForiegnKeyIndex.keys()[0]]
	invForiegnKeyClassID = invForiegnKey[1][0]
	invForiegnKeyObjectAttribute = invForiegnKey[0]

	try:
		keys = sorted(dataEntity[invForiegnKeyClassID].keys(), key=sortingFunction)
	except:
		keys = dataEntity[invForiegnKeyClassID].keys()
	rowIdx = 4
	for key in keys:
		
		for attr in dataEntity[invForiegnKeyClassID][key][1]:
			if '____' in attr:
				thisAttr = attr.split('____')[0]
				if thisAttr == invForiegnKeyObjectAttribute:

					objectID = dataEntity[invForiegnKeyClassID][key][1][attr]
					data = dataEntity[classIDCandidate[0]][objectID][0]
					objectData = dataEntity[classIDCandidate[0]][objectID][1]

					for i in invForiegnKeyIndex:
						sheet.cell(row = rowIdx, column = i+1).value = key.split('__')[1]

					for i in attributeIndex:
						value = data[attributeIndex[i][0]]
						if attributeIndex[i][1] == bool:
							value = "Yes" if value == 'true' else "No"
						sheet.cell(row = rowIdx, column = i+1).value = value

					for i in foriegnKeyIndex:
						value = objectData[foriegnKeyIndex[i][0]]
						sheet.cell(row = rowIdx, column = i+1).value = value.split('__')[1]
					rowIdx = rowIdx + 1




def mapToTable_SolidRactangle(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex):
	invForiegnKey = invForiegnKeyIndex[invForiegnKeyIndex.keys()[0]]
	invForiegnKeyClassID = invForiegnKey[1][0]
	invForiegnKeyObjectAttribute = invForiegnKey[0]
	try:
		keys = sorted(dataEntity[invForiegnKeyClassID].keys(), key=sortingFunction)
	except:
		keys = dataEntity[invForiegnKeyClassID].keys()
	rowIdx = 4
	for key in keys:
		if invForiegnKeyObjectAttribute in dataEntity[invForiegnKeyClassID][key][1]:
			objectID = dataEntity[invForiegnKeyClassID][key][1][invForiegnKeyObjectAttribute]
			data = dataEntity[classIDCandidate[0]][objectID][0]
			objectData = dataEntity[classIDCandidate[0]][objectID][1]

			# print objectID
			# print data

			for i in invForiegnKeyIndex:
				sheet.cell(row = rowIdx, column = i+1).value = key.split('__')[1]

			for i in attributeIndex:
				value = data[attributeIndex[i][0]]
				if attributeIndex[i][1] == bool:
					value = "Yes" if value == 'true' else "No"
				sheet.cell(row = rowIdx, column = i+1).value = value

			for i in foriegnKeyIndex:
				value = objectData[foriegnKeyIndex[i][0]]
				sheet.cell(row = rowIdx, column = i+1).value = value.split('__')[1]
			rowIdx = rowIdx + 1



def mapToTable(obj):
	for sheetInfo in dict_sap_to_owl:
		# Get column index of data entities
		try:
			sheet = book.get_sheet_by_name(sheetInfo['sheetName'])
		except:
			print "Spreadsheet {0} does not exist.".format(sheetInfo['sheetName'])
		else:
			classIDCandidate = sheetInfo['class']
			attributeIndex = ScanAttributes(sheetInfo['attribute'], sheet)
			foriegnKeyIndex = ScanAttributes(sheetInfo['foreignKey'], sheet)
			invForiegnKeyIndex = ScanAttributes(sheetInfo['invForeignKey'], sheet)


			if sheetInfo['sheetName'] in ['Area Auto Mesh Assignments']:
				mapToTable_AreaAutoMeshAssignments(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)
			if sheetInfo['sheetName'] in ['Area Edge Constraint Assigns']:
				mapToTable_Type1(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)
			if sheetInfo['sheetName'] in ['Area Overwrites - Joint Offsets']:
				mapToTable_AreaOverwritesJointOffsets(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)
			if sheetInfo['sheetName'] in ['MatProp 01 - General']:
				mapToTable_MatProp01General(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)
			if sheetInfo['sheetName'] in ['MatProp 02 - Basic Mech Props']:
				mapToTable_Type2(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)
			if sheetInfo['sheetName'] in ['MatProp 03a - Steel Data']:
				mapToTable_Type2(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)
			if sheetInfo['sheetName'] in ['MatProp 03b - Concrete Data']:
				mapToTable_Type2(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)
			if sheetInfo['sheetName'] in ['Area Section Properties']:
				mapToTable_AreaSectionProperties(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)
			if sheetInfo['sheetName'] in ['Area Section Assignments']:
				mapToTable_Type3(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)
			if sheetInfo['sheetName'] in ['Coordinate Systems']:
				mapToTable_CoordinateSystems(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)
			if sheetInfo['sheetName'] in ['Joint Coordinates']:
				mapToTable_JointCoordinates(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)
			if sheetInfo['sheetName'] in ['Connectivity - Area']:
				mapToTable_ConnectivityArea(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)
			if sheetInfo['sheetName'] in ['Connectivity - Frame']:
				mapToTable_ConnectivityFrame(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)
			if sheetInfo['sheetName'] in ['Frame Auto Mesh']:
				mapToTable_Type4(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)
			if sheetInfo['sheetName'] in ['Frame Releases 1 - General']:
				mapToTable_Type4(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)
			if sheetInfo['sheetName'] in ['Frame Props 01 - General']:
				mapToTable_FramePropsGeneral(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)
			if sheetInfo['sheetName'] in ['Frame Section Assignments']:
				mapToTable_FrameSectionAssignments(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)
			if sheetInfo['sheetName'] in ['Joint Restraint Assignments']:
				mapToTable_Type1(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)
			
			if sheetInfo['sheetName'] in ['SD 01 - General']:
				mapToTable_Type2(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)
			if sheetInfo['sheetName'] in ['SD 11 - Shape Plate']:
				mapToTable_ShapePlate(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)
			if sheetInfo['sheetName'] in ['SD 12 - Shape Solid Rectangle']:
				mapToTable_SolidRactangle(sheet, obj, classIDCandidate, attributeIndex, foriegnKeyIndex, invForiegnKeyIndex)






#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#
# main
#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#


if __name__ == '__main__':

	book = openpyxl.load_workbook(templateFileName)


	root = ET.parse(inputFileName).getroot()
	for obj in root.findall("{0}NamedIndividual".format(ns_owl_p)):
		objID, classID, dataProperty, objectProperty = ParseObject(obj)
		if classID in dataEntity:
			dataEntity[classID][objID] = [dataProperty, objectProperty]
		else:
			dataEntity[classID] = {}
			dataEntity[classID][objID] = [dataProperty, objectProperty]

	mapToTable(book)
	book.save(outputFileName)


	# parameter = GetParameter(root.findall('P'))
	# attribute = GetAttribute(root.items())

	
	# parameter = GetParameter(corrObj.findall('P'))

	# book = openpyxl.load_workbook(templateFileName)
	# ParseObject(root, book)


	# book.save(outputFileName)

# for a given type:
# Search all individual {ID: [a:a, b:b, ]}
# Do sorting [[ID:ID, a,a, b:b]]
# Writing to corresponding sheet.
